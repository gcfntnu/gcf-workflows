import sys
import os
import glob
import string
import re
import argparse
import copy

import six
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import peppy


METADATA_SHEET = '2. Metadata Template'
MD5_SHEET = '3. MD5 Checksums'
STUDY_START = 11
SAMPLES_START = 34
PROTOCOLS_START = 45
PAIRED_START = 65
MD5_START = 9

EXCEL_COLS = list(string.ascii_uppercase)
EXCEL_COLS.extend(['A' + i for i in string.ascii_uppercase])   

def _dedup_by_counter(x, sep='_', keep_first=False):
    """
    Make list unique by adding underscore + counter on duplicated values
    """
    index = pd.Index(x)
    if index.is_unique:
        return index
    from collections import Counter
    values = index.values.copy()
    keep = 'first' if keep_first else False
    indices_dup = index.duplicated(keep=keep)
    values_dup = values[indices_dup]
    values_set = set(values)
    counter = Counter()
    for i, v in enumerate(values_dup):
        while True:
            counter[v] += 1
            tentative_new_name = v + sep + str(counter[v])
            if tentative_new_name not in values_set:
                values_set.add(tentative_new_name)
                values_dup[i] = tentative_new_name
                break
    values[indices_dup] = values_dup
    index = pd.Index(values, name=index.name)
    return index.to_list()

    
def insert_samples(pep, workbook, supplementary_files=None, md5=None):
    df = pep.sample_table.copy()
    sheet = workbook[METADATA_SHEET]
    header = [c.value for c in sheet[SAMPLES_START-1]]
    available_raw = pd.Series(header).str.match('\*?raw file').sum()
    available_processed = pd.Series(header).str.match('\*?processed data file\.*').sum()
    n_supplementary = len(supplementary_files)
    write_new_header = False
    if supplementary_files is not None and n_supplementary >= 1:
        supplementary_files = [os.path.basename(fn) for fn in supplementary_files if fn]
        if n_supplementary > available_processed:
            for _ in range(n_supplementary - available_processed):
                header.insert(header.index('processed data file '), 'processed data file ')
            write_new_header = True
        processed_data = pd.DataFrame([supplementary_files] * df.shape[0])
        processed_data_header =  header[header.index('*processed data file ') :( header.index('*processed data file ') + n_supplementary)]
        processed_data.columns = _dedup_by_counter(processed_data_header, sep=' ', keep_first=True) # ensure unique header
        processed_data.index = df.index

    fastq_cols = [i for i in df.columns if i in ['R1', 'R2', 'I1', 'I2']]
    fastq = df[fastq_cols]
    raw_multiplier = 0
    if 'R1' in df.columns:
        len_R1 = df['R1'].apply(lambda x: len(x) if isinstance(x, (list, tuple)) else 1)
        assert all(len_R1 == len_R1[0])
        n_raw = len_R1[0]
        raw_multiplier += 1
        fastq.loc[:,'R1'] = fastq['R1'].copy().apply(lambda x: [os.path.basename(i) for i in x] if isinstance(x, (list, tuple)) else os.path.basename(x) )
    if 'R2' in df.columns:
        len_R2 = df['R2'].apply(lambda x: len(x) if isinstance(x, (list, tuple)) else 1)
        assert all(len_R2 == len_R2[0])
        assert n_raw == len_R2[0]
        raw_multiplier += 1
        fastq.loc[:,'R2'] = fastq['R2'].apply(lambda x: [os.path.basename(i) for i in x] if isinstance(x, (list, tuple)) else os.path.basename(x) )
    if 'I1' in df.columns:
        len_I1 = df['I1'].apply(lambda x: len(x) if isinstance(x, (list, tuple)) else 1)
        assert all(len_I1 == len_I1[0])
        assert n_raw == len_I1[0]
        raw_multiplier += 1
        fastq.loc[:,'I1'] = fastq['I1'].apply(lambda x: [os.path.basename(i) for i in x] if isinstance(x, (list, tuple)) else os.path.basename(x) )
    n_raw = n_raw * raw_multiplier
    if available_raw < n_raw:
        for _ in range(n_raw - available_raw):
            header.insert(header.index('raw file'), 'raw file')
        write_new_header = True
    
    # expand if needed (subsamples)
    fastq = pd.concat((pd.DataFrame(fastq[n].to_list()) for n in fastq_cols), axis=1)
    fastq_header = header[header.index('*raw file') : (header.index('*raw file') + fastq.shape[1])]
    fastq.columns = _dedup_by_counter(fastq_header, sep=' ', keep_first=True) # ensure unique header
    fastq.index = df.index
    

    df = df.rename(columns={'sample_name': '*library name', 'Sample_Group': 'treatment'})
    if 'Organism' in df.columns:
        organism = df['Organism']
    else:
        organism = df['*library name'].copy()
        organism.values[:] = pep.config.get('organism')
    df['*organism'] = organism

    title = pep.sample_table['Sample_Biosource'] + '_' + pep.sample_table['Sample_Group']
    title = title.str.replace(' ', '_')
    title = _dedup_by_counter(title, sep="_")
    df['*title'] = title

    df['*instrument model'] = pep.config['machine']

    #df['*molecule'] = pep.config['molecule']

    df['*single or paired-end'] = 'paired' if len(pep.config['read_geometry']) > 1 else 'single'

    # create room for samples
    df = pd.concat([df, processed_data, fastq], axis=1)
    n_rows = df.shape[0]
    if n_rows > 8: #available rows in template
        add_rows = n_rows - 8
        sheet.insert_rows(SAMPLES_START, amount=add_rows)

    if write_new_header:
        row_num = SAMPLES_START - 1
        for i, col_name in enumerate(header):
            col_letter = EXCEL_COLS[i]
            sheet[f'{col_letter}{row_num}'].value = col_name

    uniq_header = _dedup_by_counter(header, sep=' ', keep_first=True)
    col_map = dict(zip(uniq_header, EXCEL_COLS))
    print(df[list(set(uniq_header).intersection(df.columns))].head().T)
    for col_name in uniq_header:
        if col_name in df.columns:
            for i, val in enumerate(df[col_name].values):
                row_num = SAMPLES_START + i
                col_letter = col_map[col_name]
                sheet[f'{col_letter}{row_num}'].value = val
    return workbook

def insert_study(pep, workbook, supplementary_files=None, experimental_design=None):
    supplementary_files = copy.deepcopy(supplementary_files)
    sheet = workbook[METADATA_SHEET]
    project_id =  pep.config.get('project_id', ['NA'])[0]
    pi_filled, supp_filled = False, False
    for row in sheet.iter_rows(min_row=STUDY_START, max_col=2, max_row=STUDY_START+20, values_only=False):
        name, val = row
        #print("|{}|".format(name.value))
        if not name.value:
            if len(supplementary_files) > 0:
                val.value = os.path.basename(supplementary_files.pop(0))
                name.value = 'supplementary file'
            else:
                break
        if name.value == '*title':
            experiment_title = pep.config.get('experiment_title')
            if isinstance(experiment_title, (list, tuple)):
                experiment_title = experiment_title[0]
            experiment_title = experiment_title or project_id
            val.value = experiment_title
        elif name.value == '*summary (abstract)':
            experiment_summary = pep.config.get('experiment_summary')
            if experiment_summary in [None, '', 'NA']:
                experiment_summary = '***'
            val.value = experiment_summary
        elif name.value == 'contributor' and not pi_filled:
            PI = pep.config.get('experiment_principal_inverstigator')
            if PI:
                if PI == 'NA':
                    PI = '***'
                val.value = PI
            pi_filled = True
        elif name.value == '*experimental design':
            experimental_design = experimental_design or '***'
            val.value = experimental_design
        elif name.value == 'supplementary file' and not supp_filled:
            if len(supplementary_files) > 0:
                val.value = os.path.basename(supplementary_files.pop(0))
                supp_filled = True
    
    return workbook
    

    
def insert_protocols(pep, workbook, demultiplex_protocol=None, workflow_protocols=None, processed_data_format=None, assembly=None):
    sheet = workbook[METADATA_SHEET]
    protocol_rownum = {
        'extract_protocol': 2,
        'library_construction_protocol': 3,
        'library_strategy': 4,
        'demultiplex_protocol': 6,
        'genome_build': 11,
        'processed_data_content': 12
    }
    extraction_protocol = pep.config.get('protocols', {}).get('extraction')
    if extraction_protocol:
        row_num = PROTOCOLS_START + protocol_rownum['extract_protocol']
        sheet[f'B{row_num}'].value = extraction_protocol

    library_construction = pep.config.get('protocols', {}).get('extraction')
    if library_construction:
        row_num = PROTOCOLS_START + protocol_rownum['library_construction_protocol']
        sheet[f'B{row_num}'].value = library_construction
    
    library_strategy = pep.config.get('library_strategy')
    if library_strategy:
        row_num = PROTOCOLS_START + protocol_rownum['library_strategy']
        sheet[f'B{row_num}'].value = library_strategy
    
    if demultiplex_protocol:
        row_num = PROTOCOLS_START + protocol_rownum['demultiplex_protocol']
        sheet[f'B{row_num}'].value = demultiplex_protocol

    if workflow_protocols:
        #fixme: insert rows if number of workflow protocols > 5
        for i, (k, v) in enumerate(workflow_protocols.items()):
            row_num = PROTOCOLS_START + protocol_rownum['demultiplex_protocol'] + i
            sheet[f'B{row_num}'].value = v
    if assembly:
        row_num = PROTOCOLS_START + protocol_rownum['genome_build']
        sheet[f'B{row_num}'].value = assembly
    
    if processed_data_format:
        row_num = PROTOCOLS_START + protocol_rownum['processed_data_content']
        sheet[f'B{row_num}'].value = processed_data_format
    
    return workbook

def insert_matched_pe_fastq_merged(pep, md5, workbook):
    sheet = workbook[METADATA_SHEET]
    header = [c.value for c in sheet[PAIRED_START-1]]
    df = md5.loc[md5['sample_id'].isin(pep.sample_table.index)]
    df = df[['sample_id', 'file name']]
    row_i = 0
    for files in md5.groupby('sample_id')['file name'].agg(lambda x: x):
        if isinstance(files, six.string_types):
            files = [files]
        for col_i, val in enumerate(files):
            row_num = PAIRED_START + row_i
            col_letter = EXCEL_COLS[col_i]
            sheet[f'{col_letter}{row_num}'].value = val
        row_i += 1
    return workbook

def insert_matched_pe_fastq(pep, workbook):
    sheet = workbook[METADATA_SHEET]
    header = [c.value for c in sheet[PAIRED_START-1]]
    df = pep.sample_table.copy()
    fastq = df[[n for n in ['R1', 'R2', 'I1'] if n in df.columns]]
    fastq.columns = _dedup_by_counter(['file name']* len(fastq.columns), sep=' ')
    if fastq.shape[1] > 2:
        for i, val in enumerate(fastq.columns):
            col_letter = EXCEL_COLS[i]
            row_num = PAIRED_START - 1
            sheet[f'{col_letter}{row_num}'].value = os.path.basename(val)

    
    for i, sample_id in enumerate(fastq.index):
        sample_fastq = pd.DataFrame(fastq.loc[sample_id].to_list()).applymap(os.path.basename).T
        for row_i, row in sample_fastq.iterrows():
            for col_j, val in enumerate(row):
                row_num = PAIRED_START + i*sample_fastq.shape[0] + row_i
                col_letter = EXCEL_COLS[col_j]
                #print('{}{} > {}'.format(col_letter, row_num, val))
                sheet[f'{col_letter}{row_num}'].value = os.path.basename(val)
    return workbook
            
def read_md5(fn):
    md5 = pd.read_table(fn, sep="\s+", engine='python', names=['file checksum', 'path'])
    md5['file name'] = [os.path.basename(i) for i in md5.path]
    md5 = md5.drop('path', axis=1)
    sample_id, read_num = [], []
    for f in md5['file name']:
        m = re.match('(.*)_([RI][1-2]).fastq.gz$', f)
        if m:
            sample_id.append(m.groups()[0])
            read_num.append(m.groups()[1])
        else:
            msg = 'failed to identify sample_id from path `{}` '.format(f)
            raise ValueError(msg)
    md5['sample_id'] = sample_id
    #md5['read_num'] = read_num
    
    return md5
    

def insert_md5sums_merged(pep, workbook, md5, supplementary_files=None):
    sheet = workbook[MD5_SHEET]
    df = md5.loc[md5['sample_id'].isin(pep.sample_table.index)]
    raw_processed_index = {c.value:i for i, c in enumerate(sheet[MD5_START-2]) if c.value}
    header = ['file name', 'file checksum']
    for col_index, col_name in enumerate(header):
        for row_index, val in enumerate(df[col_name].values):
            row_index = MD5_START + row_index
            col_index = col_index + raw_processed_index['RAW FILES']
            col_letter = EXCEL_COLS[col_index]
            sheet[f'{col_letter}{row_index}'].value = val

    if supplementary_files:
        import hashlib
        for i, fn in enumerate(supplementary_files):
            if os.path.exists(fn):
                with open(fn, "rb") as fh:
                    bytes = fh.read()
                    checksum = hashlib.md5(bytes).hexdigest()
                    row = [os.path.basename(fn), checksum]
                for col_index, col_name in enumerate(header):
                    val = row[col_index]
                    col_index = col_index + raw_processed_index['PROCESSED DATA FILES']
                    col_letter = EXCEL_COLS[col_index]
                    row_index = MD5_START + i
                    sheet[f'{col_letter}{row_index}'].value = val
    return workbook

def insert_md5sums(pep, workbook, supplementary_files=None):
    sheet = workbook[MD5_SHEET]
    raw_processed_index = {c.value:i for i, c in enumerate(sheet[MD5_START-2]) if c.value}
    row_i = 0
    for read_name in ['R1', 'R2', 'I1']:
        if read_name in pep.sample_table:
            read_md5_name = '{}_md5sum'.format(read_name)
            for file_name, file_checksum in pep.sample_table[[read_name, read_md5_name]].values:
                if isinstance(file_name, six.string_types):
                    file_name = [file_name]
                    file_checksum = [file_checksum]
                for r, m in zip(file_name, file_checksum):
                    fn = os.path.basename(r)
                    row_index = MD5_START + row_i
                    col_letter = EXCEL_COLS[raw_processed_index['RAW FILES']]
                    #print('{}{} > {}'.format(col_letter, row_index, fn))
                    sheet[f'{col_letter}{row_index}'].value = fn
                    col_letter = EXCEL_COLS[raw_processed_index['RAW FILES'] + 1]
                    sheet[f'{col_letter}{row_index}'].value = m
                    row_i += 1
    
    if supplementary_files:
        import hashlib
        for i, fn in enumerate(supplementary_files):
            if os.path.exists(fn):
                with open(fn, "rb") as fh:
                    bytes = fh.read()
                    checksum = hashlib.md5(bytes).hexdigest()
                fn = os.path.basename(fn)
                row_index = MD5_START + i
                col_letter = EXCEL_COLS[raw_processed_index['PROCESSED DATA FILES']]
                #print('{}{} > {}'.format(col_letter, row_index, fn))
                sheet[f'{col_letter}{row_index}'].value = fn
                col_letter = EXCEL_COLS[raw_processed_index['PROCESSED DATA FILES'] + 1]
                sheet[f'{col_letter}{row_index}'].value = checksum
    
    return workbook


def get_workflow_protocols(repo_dir, workflow):
    protocols = {}
    with open(os.path.join(repo_dir, ".git", "HEAD"), 'r') as head_fh:
        branch = head_fh.read().split('/')[-1].rstrip()
    with open(os.path.join(repo_dir, ".git", "refs", "heads", branch), 'r') as commit_fh:
        commit = commit_fh.read().rstrip()
    workflow_version = "github.com/gcfntnu/gcf-workflows/tree/{} commit {}".format(branch, commit)
    protocols['gcf-workflow'] = f'FASTQ files were processsed by the {workflow} pipline of gcf-tools ({workflow_version}) ' 
    return protocols

def create_parser():
    parser = argparse.ArgumentParser(formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    parser.add_argument("-S", "--sample-info", required=True, 
                        help="Path to sample meta data")
    
    parser.add_argument("-T", "--template", required=True, 
                        help="Path to GEO seq_template.xlsx")

    parser.add_argument("-P", "--pep", type=argparse.FileType('r'), required=True,
                        help="Path to peppy project")

    parser.add_argument("-A", "--assembly",
                        help="", default=None)

    parser.add_argument("--read-geometry",
                        help="", required=True)

    parser.add_argument("--md5sum-file",
                        help="", required=False)

    parser.add_argument("--processed-data", default=None, nargs='+',
                        help="")

    parser.add_argument("--demultiplex-protocol", default=None,
                        help="")

    parser.add_argument("--repo-dir", default=None,
                        help="")
    
    parser.add_argument("--workflow", default=None, 
                        help="workflow name")

    parser.add_argument("-o", "--output", required=True,
                        help="output excel file")
    return parser

if __name__ == '__main__':
    args = create_parser().parse_args()
    pep = peppy.Project(args.pep.name)
    merged_fastq_output = False 
    if args.md5sum_file:
        md5 = read_md5(args.md5sum_file)
        merged_fastq_output = True
    experimental_design = pep.config.get('experimental_design', '***')
    supplementary_files = args.processed_data
    if supplementary_files is not None:
        description_file = 'tab separated file with header and row names in first column' 
        protocol_file = ''
    demultiplex_protocol = args.demultiplex_protocol
    if demultiplex_protocol is not None:
        with open(demultiplex_protocol) as fh:
            demultiplex_protocol = fh.read().splitlines()
    else:
        demultiplex_protocol = "Sequencing basecalls were demultiplexed and converted into FASTQ using bcl2fastq v2.20.0.422"
    workflow_protocols = get_workflow_protocols(args.repo_dir, args.workflow)
    processed_data_format = None
    if processed_data_format is None:
        processed_data_format = "Tab seprated files with header and rownames in first column"
    if args.assembly == 'NA':
        assembly = None
    else:
        assembly = args.assembly
    
    wb = load_workbook(args.template)
    wb = insert_study(pep, wb, experimental_design=experimental_design, supplementary_files=supplementary_files)
    wb = insert_samples(pep, wb, supplementary_files=supplementary_files)
    wb = insert_protocols(pep, wb, demultiplex_protocol, workflow_protocols, processed_data_format, assembly)
    if len(pep.config['read_geometry']) > 1:
        if merged_fastq_output:
            insert_matched_pe_fastq_merged(pep, md5, wb)
        else:
            wb = insert_matched_pe_fastq(pep, wb)
            
    wb = insert_md5sums(pep, wb, supplementary_files=supplementary_files)

    wb.save(args.output)
    
    
